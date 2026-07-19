"""
Generic tabular (CSV/XLSX) collector. Used for sources that publish their
breach list as a downloadable dataset (HHS OCR portal export, Mass.gov
annual reports, Washington AG dataset, Privacy Rights Clearinghouse
chronology). XLSX support is included since several state AG offices
prefer Excel exports over CSV.
"""
from __future__ import annotations

import abc
import csv
import io

import openpyxl

from app.collectors.base import BaseCollector, NormalizedRecord


class TabularCollector(BaseCollector):
    default_document_type = "ag_notification_letter"
    file_kind: str = "csv"  # or "xlsx"

    async def fetch_raw(self) -> bytes:
        # FIX: guard against feed_url being NULL in the DB (e.g. HHS OCR portal
        # whose CSV export URL was not discovered at seed time). Return empty
        # bytes; parse() below handles the empty case cleanly instead of
        # crashing with TypeError: Invalid type for url ... got NoneType.
        if not self.source_row.feed_url:
            return b""
        resp = await self.client.get(self.source_row.feed_url)
        resp.raise_for_status()
        return resp.content

    def rows(self, raw: bytes) -> list[dict]:
        if self.file_kind == "xlsx":
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            out = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                out.append(dict(zip(header, row)))
            return out
        text = raw.decode("utf-8", errors="ignore")
        return list(csv.DictReader(io.StringIO(text)))

    @abc.abstractmethod
    def map_row(self, row: dict) -> NormalizedRecord | None: ...

    async def parse(self, raw: bytes) -> list[NormalizedRecord]:
        # FIX: return early on empty bytes (e.g. when feed_url is None and
        # fetch_raw() returned b"") rather than trying to parse an empty CSV.
        if not raw:
            return []
        out = []
        for row in self.rows(raw):
            rec = self.map_row(row)
            if rec:
                out.append(rec)
        return out
